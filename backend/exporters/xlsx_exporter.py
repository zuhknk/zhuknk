"""Excel 导出"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_xlsx(data: dict) -> bytes:
    """将分析结果导出为 Excel"""
    wb = Workbook()

    reviews = data.get("reviews", [])
    findings = data.get("findings", [])
    requirements = data.get("requirements", [])
    test_cases = data.get("test_cases", [])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e2d3d", end_color="1e2d3d", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 概览 Sheet
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    ws.append(["Total Reviews", len(reviews)])
    ws.append(["Findings", len(findings)])
    ws.append(["Requirements", len(requirements)])
    ws.append(["Test Cases", len(test_cases)])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill

    # 评论 Sheet
    if reviews:
        ws = wb.create_sheet("Reviews")
        headers = ["ID", "Rating", "Title", "Content", "Author", "Version", "Date"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in reviews:
            ws.append([
                r.get("review_id", ""),
                r.get("rating", 0),
                r.get("title", ""),
                r.get("content", ""),
                r.get("author", ""),
                r.get("version", ""),
                r.get("date", ""),
            ])

    # 发现 Sheet
    if findings:
        ws = wb.create_sheet("Findings")
        headers = ["ID", "Topic", "Severity", "Sentiment", "Description", "Confidence", "Samples"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for f in findings:
            ws.append([
                f.get("finding_id", ""),
                f.get("topic", ""),
                f.get("severity", ""),
                f.get("sentiment", ""),
                f.get("description", ""),
                f.get("confidence", 0),
                f.get("sample_count", 0),
            ])
            # 颜色标记严重程度
            row_idx = ws.max_row
            severity = f.get("severity", "")
            color_map = {"critical": "FF5252", "high": "FF9100", "medium": "FFAB00", "low": "448AFF"}
            if severity in color_map:
                ws.cell(row=row_idx, column=3).fill = PatternFill(
                    start_color=color_map[severity], end_color=color_map[severity], fill_type="solid"
                )

    # 需求 Sheet
    if requirements:
        ws = wb.create_sheet("Requirements")
        headers = ["ID", "Title", "Priority", "Description", "User Story"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in requirements:
            ws.append([
                r.get("req_id", ""),
                r.get("title", ""),
                r.get("priority", ""),
                r.get("description", ""),
                r.get("user_story", ""),
            ])

    # 测试用例 Sheet
    if test_cases:
        ws = wb.create_sheet("Test Cases")
        headers = ["ID", "Req ID", "Title", "Priority", "Steps", "Expected Result"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for tc in test_cases:
            ws.append([
                tc.get("case_id", ""),
                tc.get("req_id", ""),
                tc.get("title", ""),
                tc.get("priority", ""),
                "\n".join(tc.get("steps", [])),
                tc.get("expected_result", ""),
            ])

    # 自动调整列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
